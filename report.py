from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference, Series
from openpyxl.styles import Font, Border, Side
import os

from chatbot import chat_with_data

# Define a border style
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
def write_data_to_sheet(ws, data, start_row, start_col, title):
    # Add the dataset header
    ws.cell(row=start_row - 2, column=start_col, value=title).font = Font(bold=True)
    
    # Write the data to the specified starting cell
    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Make the header row and index column bold
            if r_idx == start_row or c_idx == start_col:
                cell.font = Font(bold=True)
            # Add borders to the cell
            cell.border = thin_border
            
def create_bar_chart(ws, data, title, x_axis_title, y_axis_title, start_row, start_col):
    write_data_to_sheet(ws, data, start_row, start_col, title)

    chart = BarChart()
    data_ref = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(data), max_col=start_col + len(data.columns) - 1)
    categories = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + len(data))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.style = 10  # Standard color style
    chart.y_axis.majorGridlines = None

    # Ensure the axis titles are visible
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    
    return chart, f"G{start_row - 2}"

def create_pie_chart(ws, data, title, start_row, start_col):
    write_data_to_sheet(ws, data, start_row, start_col, title)

    chart = PieChart()
    data_ref = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(data), max_col=start_col + len(data.columns) - 1)
    categories = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + len(data))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = title
    chart.style = 10  # Standard color style
    
    return chart, f"G{start_row - 2}"

def create_line_chart(ws, data, title, x_axis_title, y_axis_title, start_row, start_col):
    write_data_to_sheet(ws, data, start_row, start_col, title)

    chart = LineChart()
    data_ref = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(data), max_col=start_col + len(data.columns) - 1)
    categories = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + len(data))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.style = 10  # Standard color style
    chart.y_axis.majorGridlines = None

    # Ensure the axis titles are visible
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    
    return chart, f"G{start_row - 2}"


def create_report(dataset,prompt,folder):
    df = dataset
    
    # Create a new Excel workbook and add worksheets
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"
    
    # Write the original DataFrame to the Data worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    # Add a new worksheet for the report
    ws_report = wb.create_sheet(title="Report")
    
    # Add a title to the Report worksheet
    ws_report.merge_cells('A1:I1')
    ws_report['A1'] = "Analysis Report"
    ws_report['A1'].style = 'Title'
    
    chart_configurations = prompt
    
    start_row = 5
    start_col = 1
    chart_functions = {
    'bar': create_bar_chart,
    'pie': create_pie_chart,
    'line': create_line_chart
    }
    
    # Loop through the chart configurations and create charts
    for config in chart_configurations:
        chart_func = chart_functions[config['type']]
        if config['type'] == 'bar' or config['type'] == 'line':
            data = chat_with_data(df, config['prompt'], folder)
            # print(data)
            chart, chart_cell = chart_func(ws_report, data, config['title'], config['x_axis_title'], config['y_axis_title'], start_row, start_col)
        else:
            data = chat_with_data(df, config['prompt'], folder)
            # print(data)
            chart, chart_cell = chart_func(ws_report, data, config['title'], start_row, start_col)
        
        # Add the chart to the worksheet
        ws_report.add_chart(chart, chart_cell)
        
        # Update start_row for the next chart
        start_row += (len(data) if len(data) > 15 else 15) + 4
    
    # Remove gridlines from the 'Report' worksheet
    ws_report.sheet_view.showGridLines = False
    
    # Save the workbook
    output_file = f'report/{folder}/formatted_output.xlsx'
    
    # Ensure the directory exists
    output_dir = os.path.dirname(output_file)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    wb.save(output_file)
    
    return output_file
    
    
    