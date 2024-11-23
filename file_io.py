import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
import os

def save_to_excel(df, folder):
    output_file = f'file/{folder}/formatted_output.xlsx'
    
    # Ensure the directory exists
    output_dir = os.path.dirname(output_file)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    # Save pivot_table to Excel with custom formatting
    pivot_table = df
    # Initialize an Excel writer using openpyxl
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Write the DataFrame to Excel
        pivot_table.to_excel(writer, sheet_name='Data', startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['Data']
        
        # Apply styling and formatting
        header_font = Font(bold=True)
        for cell in worksheet[2]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Set alignment for the data rows
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
        
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
            worksheet.column_dimensions[col[0].column_letter].width = max_length
        
    return output_file